from odoo import models, fields, api
import openpyxl
import base64
from io import BytesIO
from odoo.exceptions import UserError, ValidationError

class ResPartnerBank(models.Model):
    _inherit = 'res.partner.bank'

    _sql_constraints = [(
        'unique_number',
        'unique (1=1)',
        'The combination Account Number/Partner must be unique.'
    )]


class IBANWIZ(models.TransientModel):
    _name = 'wiz.update.iban'
    _description = 'Update Partner Iban'

    excel_file = fields.Binary(string="Excel File")

    def read_excel_file(self):
        self.ensure_one()
        if not self.excel_file:
            raise UserError("No Excel file uploaded.")
        file_data = base64.b64decode(self.excel_file)
        file_stream = BytesIO(file_data)
        workbook = openpyxl.load_workbook(filename=file_stream)
        data = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            print(f"Reading data from sheet: {sheet_name}")

            # Iterate through the rows and read the data from columns B and C (indexes 1 and 2)
            for row in sheet.iter_rows(min_row=1, values_only=True):  # Start from row 1
                column_b = row[1] if len(row) > 1 else None  # Column B (index 1)
                column_c = row[5] if len(row) > 5 else None  # Column C (index 2)

                # Append the data if either column B or C has a value
                if column_b is not None or column_c is not None:
                    data.append((column_b, column_c))  # Include sheet name for reference

            res_bank = self.env['res.partner.bank'].search([])
            for line in res_bank:
               for account, beneficiary_name in data:
                    print(beneficiary_name,line.partner_id.employee_ids)
                    if any(employee.registration_number == beneficiary_name for employee in
                           line.partner_id.employee_ids):
                        line.write({'acc_number': account})
                    else:
                      print("No matching record found for beneficiary name:", beneficiary_name)
